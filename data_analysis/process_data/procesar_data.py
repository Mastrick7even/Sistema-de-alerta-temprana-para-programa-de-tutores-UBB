import pandas as pd
import openpyxl
import glob
import os
import re
import warnings

warnings.filterwarnings('ignore', category=UserWarning)

# --- CONFIGURACIÓN ---
CARPETA_DATOS = "raw_data"
ARCHIVO_SALIDA = "dataset_sat_unificado.csv"

# Hojas a ignorar (Tablas dinámicas, instrucciones, etc)
HOJAS_IGNORAR = [
    "instrucc", "resumen", "hoja1", "base", "ejemplo", "decálogo", "decalogo", 
    "ruta", "contacto", "protocolo", "derivación", "actividades", "tabla"
]

def es_rut_chileno(texto):
    """Detecta si un string parece un RUT chileno (con o sin puntos)"""
    if not texto: return False
    # Limpieza: quitar puntos y espacios
    limpio = str(texto).replace('.', '').strip().upper()
    # Regex: 7 o 8 digitos + guion (opcional) + K o numero
    # Ej: 12345678-K, 12345678K, 123456789
    patron = r'^\d{7,8}-?[\dK]$'
    return bool(re.match(patron, limpio))

def detectar_columnas_por_contenido(ws):
    """
    Escanea las primeras 20 filas buscando columnas que parezcan RUTs.
    Retorna el índice de la columna RUT y Nombre estimado.
    """
    max_filas_a_escanear = 30
    
    # Diccionario para contar aciertos por columna
    # { indice_columna: cantidad_de_ruts_encontrados }
    col_scores_rut = {}
    
    # Iteramos filas buscando patrones
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_filas_a_escanear, values_only=True)):
        for col_idx, valor in enumerate(row):
            if not valor: continue
            val_str = str(valor).strip()
            
            # Chequear si parece RUT
            if es_rut_chileno(val_str):
                col_scores_rut[col_idx] = col_scores_rut.get(col_idx, 0) + 1

    # Decisión: ¿Cuál columna ganó?
    idx_rut = -1
    if col_scores_rut:
        # La columna con más aciertos de RUT
        idx_rut = max(col_scores_rut, key=col_scores_rut.get)
        # Filtro: Debe haber al menos 2 RUTs detectados para confiar
        if col_scores_rut[idx_rut] < 2:
            idx_rut = -1

    # Si encontramos RUT, asumimos que el Nombre está a la derecha (+1) o (+2)
    # y buscamos dónde empiezan las alertas (colores).
    idx_nombre = -1
    fila_inicio_datos = -1
    
    if idx_rut != -1:
        idx_nombre = idx_rut + 1 # Por defecto
        
        # Ahora necesitamos saber en qué fila empiezan los datos reales
        # Buscamos la primera fila donde la columna RUT tenga un RUT válido
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_filas_a_escanear, values_only=True)):
             if row[idx_rut] and es_rut_chileno(row[idx_rut]):
                 fila_inicio_datos = i + 1 # Base 1 para openpyxl
                 break

    return idx_rut, idx_nombre, fila_inicio_datos

def obtener_riesgo_color(celda):
    """Devuelve riesgo (0, 2, 3) basado en color Hex"""
    if celda.fill and celda.fill.start_color:
        try:
            if not celda.fill.start_color.index: return 0
            # Convertimos a string y mayusculas
            c = str(celda.fill.start_color.index).upper()
            
            # Filtro colores raros o theme colors (suelen ser ints)
            if len(c) < 3: return 0 

            # Mapeo de Colores (Ajustar según tus excels)
            # Rojos
            if c in ['FFFF0000', '00FF0000', 'RED', 'FFFFC7CE']: return 3
            # Amarillos
            if c in ['FFFFFF00', 'YELLOW', 'FFFFEB9C']: return 2
            # Blanco (solo si quieres trackear ghosting explícito)
            # if c == 'FFFFFFFF': return -1 
        except:
            pass
    return 0

def procesar_excels():
    data_final = []
    archivos = glob.glob(os.path.join(CARPETA_DATOS, "*.xlsx"))
    
    if not archivos:
        print("❌ No hay archivos .xlsx")
        return

    print(f"📂 Encontrados {len(archivos)} archivos. Iniciando Escaneo Inteligente...")

    for archivo in archivos:
        nombre_archivo = os.path.basename(archivo)
        print(f"\n📄 {nombre_archivo}")
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            
            # Año default
            anio_match = re.search(r'20\d{2}', nombre_archivo)
            anio = anio_match.group(0) if anio_match else "2024"

            for sheet_name in wb.sheetnames:
                if any(x in sheet_name.lower() for x in HOJAS_IGNORAR): continue
                ws = wb[sheet_name]
                
                # --- DETECCIÓN INTELIGENTE ---
                idx_rut, idx_nombre, fila_data = detectar_columnas_por_contenido(ws)
                
                if idx_rut == -1 or fila_data == -1:
                    # print(f"   ⏩ Saltando hoja '{sheet_name}': No se detectaron RUTs válidos.")
                    continue
                
                print(f"   ✅ Hoja '{sheet_name}': Datos detectados desde Fila {fila_data}. (RUT en Col {idx_rut})")

                # Procesar Filas
                count = 0
                for row in ws.iter_rows(min_row=fila_data):
                    # Evitar desbordamiento
                    if idx_rut >= len(row): continue
                    
                    val_rut = str(row[idx_rut].value).strip() if row[idx_rut].value else ""
                    
                    # Doble chequeo: Es RUT?
                    if not es_rut_chileno(val_rut): continue

                    # Extraer nombre (si existe columna)
                    val_nombre = "Desconocido"
                    if idx_nombre < len(row) and row[idx_nombre].value:
                        val_nombre = str(row[idx_nombre].value).strip()
                        # Si el nombre parece un email, es que la columna está corrida, limpiamos
                        if "@" in val_nombre: val_nombre = "Error Columna"

                    # Escanear COLORES (Alertas)
                    # Escaneamos TODA la fila excepto RUT y Nombre para buscar colores rojos/amarillos
                    rojos = 0
                    amarillos = 0
                    obs_list = []

                    for i, celda in enumerate(row):
                        if i == idx_rut or i == idx_nombre: continue # Saltamos info personal
                        
                        riesgo = obtener_riesgo_color(celda)
                        if riesgo == 3: rojos += 1
                        if riesgo == 2: amarillos += 1
                        
                        # Guardar texto de celdas coloreadas o lejanas (posibles observaciones)
                        # OJO: Solo guardamos texto si la celda tiene COLOR o es muy larga (obs espontanea)
                        if (riesgo > 0 or (celda.value and len(str(celda.value)) > 15)) and celda.value:
                            txt = str(celda.value).strip()
                            if txt.lower() not in ['si', 'no', 'n/a']:
                                obs_list.append(txt)

                    data_final.append({
                        'archivo': nombre_archivo,
                        'anio': anio,
                        'tutor': sheet_name,
                        'rut': val_rut,
                        'nombre': val_nombre,
                        'cant_rojos': rojos,
                        'cant_amarillos': amarillos,
                        'observaciones': " || ".join(set(obs_list)) # set elimina duplicados
                    })
                    count += 1
                
                print(f"      -> {count} alumnos extraídos.")

        except Exception as e:
            print(f"   ❌ Error leyendo {nombre_archivo}: {e}")

    # Guardar
    if data_final:
        df = pd.DataFrame(data_final)
        # Limpieza final de RUTs duplicados (nos quedamos con el peor caso de riesgo si se repite en el mismo año)
        # Opcional: df = df.sort_values('cant_rojos', ascending=False).drop_duplicates(subset=['rut', 'anio'])
        
        print(f"\n✨ GENERACIÓN EXITOSA: {len(df)} registros totales.")
        print(df.head())
        df.to_csv(ARCHIVO_SALIDA, index=False, encoding='utf-8-sig')
    else:
        print("\n⚠️ No se generaron datos.")

if __name__ == "__main__":
    procesar_excels()