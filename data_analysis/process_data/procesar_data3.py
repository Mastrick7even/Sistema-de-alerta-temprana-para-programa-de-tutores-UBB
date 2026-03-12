import pandas as pd
import openpyxl
import glob
import os
import re
import warnings

warnings.filterwarnings('ignore', category=UserWarning)

# --- CONFIGURACIÓN ---
CARPETA_DATOS = "raw_data"
ARCHIVO_SALIDA = "bitacora_estructurada_final.csv"

# Hojas a ignorar
HOJAS_IGNORAR = [
    "instrucc", "resumen", "hoja1", "base", "ejemplo", "decálogo", "decalogo", 
    "ruta", "contacto", "protocolo", "derivación", "actividades", "tabla", "validac", "listado"
]

# Mapa de sinónimos para encabezados
MAPA_COLUMNAS_BASE = {
    "Teléfono": ["telefono", "fono", "celular", "movil", "contact"],
    "Correo": ["correo", "mail", "email"],
    "Lugar Procedencia": ["procedencia", "lugar", "ciudad", "origen", "comuna"],
    "Grupo Familiar": ["familiar", "familia", "vive con"],
    "Beneficio": ["beneficio", "gratuidad", "beca", "arancel", "financiamiento"],
    "Observaciones": ["observacion", "observaciones", "comentario", "detalle", "info adicional"]
}

def limpiar_texto(val):
    if not val: return ""
    return str(val).replace('\n', ' ').strip()

def es_rut_chileno(texto):
    """Detecta patrón de RUT (flexible)"""
    if not texto: return False
    limpio = str(texto).replace('.', '').strip().upper()
    # Al menos 6 digitos, un guion y digito verificador
    return len(limpio) > 5 and any(char.isdigit() for char in limpio) and '-' in limpio

def detectar_riesgo_etiqueta(celda):
    """Retorna [ROJO] o [AMARILLO] si la celda tiene color"""
    if celda.fill and celda.fill.start_color:
        try:
            if not celda.fill.start_color.index: return ""
            c = str(celda.fill.start_color.index).upper()
            if len(c) < 3: return "" 
            
            if c in ['FFFF0000', '00FF0000', 'RED', 'FFFFC7CE', '3']: return "[ROJO] "
            if c in ['FFFFFF00', 'YELLOW', 'FFFFEB9C', '2']: return "[AMARILLO] "
        except:
            pass
    return ""

def escanear_ubicacion_datos(ws):
    """
    Método Sherlock: Escanea la hoja buscando la columna con más RUTs
    Retorna: (indice_columna_rut, indice_fila_inicio_datos)
    """
    max_filas = 50
    scores_rut = {} # {col_idx: cantidad_ruts}
    
    # Muestreo para encontrar la columna RUT
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_filas, values_only=True)):
        for col_idx, valor in enumerate(row):
            if es_rut_chileno(str(valor)):
                scores_rut[col_idx] = scores_rut.get(col_idx, 0) + 1
    
    if not scores_rut:
        return -1, -1

    # Ganador: La columna con más RUTs
    idx_rut = max(scores_rut, key=scores_rut.get)
    
    # Encontrar la primera fila que tenga un RUT en esa columna
    fila_inicio = -1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_filas, values_only=True)):
        if es_rut_chileno(str(row[idx_rut])):
            fila_inicio = i + 1 # Base 1
            break
            
    return idx_rut, fila_inicio

def generar_mapa_columnas(ws, fila_datos, idx_rut):
    """
    Intenta adivinar qué es cada columna mirando la fila ANTERIOR a los datos.
    Si no encuentra encabezados, devuelve un mapa básico solo con RUT.
    """
    mapa = {}
    
    # Asumimos que los headers están 1 fila antes de los datos
    fila_header = fila_datos - 1
    if fila_header < 1: return { 'rut': idx_rut } # Sin headers posibles

    # Leemos la fila de headers
    headers = []
    # Truco: leemos toda la fila, asumiendo un ancho máximo razonable (ej: 40 cols)
    row_cells = list(ws.iter_rows(min_row=fila_header, max_row=fila_header, max_col=40))[0]
    
    for i, cell in enumerate(row_cells):
        val = str(cell.value).lower().strip() if cell.value else ""
        headers.append(val)
        
        # Mapeo de RUT (Confirmación)
        if i == idx_rut:
            mapa['rut'] = i
            continue
            
        # 1. Buscar Columnas Base (Teléfono, etc.)
        encontrado = False
        for nombre_meta, sinonimos in MAPA_COLUMNAS_BASE.items():
            if any(s in val for s in sinonimos) and "alerta" not in val:
                mapa[nombre_meta] = i
                encontrado = True
                break
        if encontrado: continue

        # 2. Buscar Alertas Numeradas
        match_num = re.search(r'(\d+)[°º]?\s*alerta|alerta\s*(\d+)', val)
        if match_num:
            # Obtener el número de cualquiera de los dos grupos de captura
            num = match_num.group(1) if match_num.group(1) else match_num.group(2)
            mapa[f"Alerta {num}"] = i
            continue

        # 3. Alerta Espontánea
        if "espontanea" in val or "espontánea" in val:
            if "motivo" in val: mapa["Alerta Espontánea Motivo"] = i
            elif "detalle" in val or "caso" in val: mapa["Alerta Espontánea Detalle"] = i
            else: 
                if "Alerta Espontánea Motivo" not in mapa:
                    mapa["Alerta Espontánea Motivo"] = i

    # Si no detectamos columna Nombre explícita, asumimos RUT + 1
    if 'nombre' not in mapa and 'rut' in mapa:
         # Verificamos que rut+1 no sea una alerta mapeada
         idx_nombre_cand = mapa['rut'] + 1
         if idx_nombre_cand not in mapa.values():
             mapa['nombre'] = idx_nombre_cand

    return mapa

def procesar_excels():
    data_global = []
    archivos = glob.glob(os.path.join(CARPETA_DATOS, "*.xlsx"))
    
    print(f"📂 Procesando {len(archivos)} archivos (Modo Híbrido)...")

    for archivo in archivos:
        nombre_archivo = os.path.basename(archivo)
        # print(f"📄 {nombre_archivo}")
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            anio_match = re.search(r'20\d{2}', nombre_archivo)
            anio = anio_match.group(0) if anio_match else "2024"

            for sheet_name in wb.sheetnames:
                if any(ign in sheet_name.lower() for ign in HOJAS_IGNORAR): continue
                ws = wb[sheet_name]
                
                # 1. ENCONTRAR DATOS (Método Sherlock)
                idx_rut, fila_datos = escanear_ubicacion_datos(ws)
                
                if idx_rut == -1:
                    continue # Hoja vacía o sin RUTs

                # 2. MAPEAR COLUMNAS (Método Estructurado)
                mapa = generar_mapa_columnas(ws, fila_datos, idx_rut)
                
                # Invertir mapa para saber qué columnas ya tienen dueño { indice: 'NombreCampo' }
                cols_ocupadas = {v: k for k, v in mapa.items()}
                
                # 3. EXTRAER FILAS
                count_sheet = 0
                for row in ws.iter_rows(min_row=fila_datos):
                    if idx_rut >= len(row): continue
                    
                    celda_rut = row[idx_rut]
                    val_rut = limpiar_texto(celda_rut.value)
                    if not es_rut_chileno(val_rut): continue

                    # Estructura base
                    item = {
                        "Origen": nombre_archivo, "Anio": anio, "Tutor": sheet_name,
                        "rut": val_rut, "nombre": ""
                    }
                    
                    # Extraer Nombre (si existe mapeo)
                    if 'nombre' in mapa and mapa['nombre'] < len(row):
                        val_nom = limpiar_texto(row[mapa['nombre']].value)
                        # Filtro simple: si tiene números (teléfono), lo ignoramos
                        if not any(c.isdigit() for c in val_nom):
                            item['nombre'] = val_nom

                    # Extraer Campos Mapeados (Teléfono, Alerta 1, etc)
                    for campo, idx in mapa.items():
                        if campo in ['rut', 'nombre']: continue
                        if idx < len(row):
                            celda = row[idx]
                            txt = limpiar_texto(celda.value)
                            tag = detecting_riesgo = detectar_riesgo_etiqueta(celda)
                            
                            if txt or tag:
                                item[campo] = f"{tag}{txt}"

                    # 4. RED DE SEGURIDAD (Capturar alertas no mapeadas)
                    # Escaneamos celdas que NO están en el mapa pero tienen COLOR
                    obs_extra = []
                    for i, celda in enumerate(row):
                        if i in cols_ocupadas: continue # Ya la leímos
                        
                        tag = detectar_riesgo_etiqueta(celda)
                        txt = limpiar_texto(celda.value)
                        
                        # Si tiene color O es un texto largo (posible observación)
                        es_texto_largo = len(txt) > 15 and "si" not in txt.lower() and "no" not in txt.lower()
                        
                        if tag or (es_texto_largo and i > idx_rut): # Solo a la derecha del RUT
                            contenido = f"{tag}{txt}"
                            if contenido.strip():
                                obs_extra.append(contenido)
                    
                    if obs_extra:
                        # Concatenamos a Observaciones si ya existe, o creamos
                        existente = item.get("Observaciones", "")
                        nuevo = " || ".join(obs_extra)
                        item["Observaciones"] = f"{existente} || {nuevo}" if existente else nuevo

                    data_global.append(item)
                    count_sheet += 1
                
                # print(f"   ✅ {sheet_name}: {count_sheet} regs")

        except Exception as e:
            print(f"❌ Error en {nombre_archivo}: {e}")

    # 5. EXPORTAR
    if data_global:
        df = pd.DataFrame(data_global)
        
        # Ordenar columnas
        cols_fijas = ["Origen", "Anio", "Tutor", "rut", "nombre", "Teléfono", "Correo", 
                      "Lugar Procedencia", "Grupo Familiar", "Beneficio", "Observaciones"]
        
        # Ordenar Alertas
        cols_dinamicas = [c for c in df.columns if c not in cols_fijas]
        def sort_alertas(c):
            nums = re.findall(r'\d+', c)
            return int(nums[0]) if nums else 999
            
        cols_alertas = sorted([c for c in cols_dinamicas if "Alerta" in c and "Espontánea" not in c], key=sort_alertas)
        cols_resto = [c for c in cols_dinamicas if c not in cols_alertas]
        
        df = df[cols_fijas + cols_alertas + cols_resto]
        df.fillna("", inplace=True)
        
        print(f"\n✨ PROCESO TERMINADO. Total: {len(df)} registros recuperados.")
        df.to_csv(ARCHIVO_SALIDA, index=False, encoding='utf-8-sig')
    else:
        print("⚠️ No se encontraron datos.")

if __name__ == "__main__":
    procesar_excels()