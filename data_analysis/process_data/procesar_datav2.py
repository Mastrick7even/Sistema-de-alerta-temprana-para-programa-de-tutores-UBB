import pandas as pd
import openpyxl
import glob
import os
import re
import warnings

warnings.filterwarnings('ignore', category=UserWarning)

# --- CONFIGURACIÓN ---
CARPETA_DATOS = "raw_data"
ARCHIVO_SALIDA = "dataset_sat_unificado.csv"

# Hojas a ignorar (Tablas dinámicas, instrucciones, etc)
HOJAS_IGNORAR = [
    "instrucc", "resumen", "hoja1", "base", "ejemplo", "decálogo", "decalogo", 
    "ruta", "contacto", "protocolo", "derivación", "actividades", "tabla"
]

def es_rut_chileno(texto):
    """Detecta si un string parece un RUT chileno"""
    if not texto: return False
    limpio = str(texto).replace('.', '').strip().upper()
    # Debe tener al menos 6 caracteres y contener algún dígito
    # Regex relajado: numeros + guion opcional + numero/k
    return len(limpio) > 5 and any(char.isdigit() for char in limpio) and '-' in limpio

def es_nombre_valido(texto):
    """
    Verifica si el texto es un nombre de persona válido.
    Retorna True si es texto puro.
    Retorna False si parece telefono, email, fecha o vacio.
    """
    if not texto: return False
    txt = str(texto).strip()
    
    # 1. Longitud mínima
    if len(txt) < 3: return False
    
    # 2. Si contiene números, NO es un nombre (es telefono, fecha o rut mal puesto)
    if any(char.isdigit() for char in txt): return False
    
    # 3. Si tiene @, es un correo
    if '@' in txt: return False
    
    # 4. Palabras clave a ignorar
    if txt.lower() in ["si", "no", "n/a", "pendiente", "observacion", "sin alerta"]: return False
    
    return True

def obtener_riesgo_color(celda):
    """Devuelve riesgo (0, 2, 3) basado en color Hex"""
    if celda.fill and celda.fill.start_color:
        try:
            if not celda.fill.start_color.index: return 0
            c = str(celda.fill.start_color.index).upper()
            if len(c) < 3: return 0 
            # Rojos
            if c in ['FFFF0000', '00FF0000', 'RED', 'FFFFC7CE', '3']: return 3
            # Amarillos
            if c in ['FFFFFF00', 'YELLOW', 'FFFFEB9C', '2']: return 2
        except:
            pass
    return 0

def buscar_columna_rut(ws):
    """Escanea las primeras filas para encontrar el índice de la columna RUT"""
    max_filas = 30
    scores_rut = {} # {col_idx: cantidad_ruts}
    
    for row in ws.iter_rows(min_row=1, max_row=max_filas, values_only=True):
        for col_idx, valor in enumerate(row):
            if es_rut_chileno(str(valor)):
                scores_rut[col_idx] = scores_rut.get(col_idx, 0) + 1
                
    if scores_rut:
        # Devolver la columna con más RUTs encontrados
        return max(scores_rut, key=scores_rut.get)
    return -1

def encontrar_inicio_datos(ws, col_rut):
    """Encuentra en qué fila empiezan los datos reales"""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if es_rut_chileno(str(row[col_rut])):
            return i + 1 # Base 1
    return -1

def procesar_excels():
    data_final = []
    archivos = glob.glob(os.path.join(CARPETA_DATOS, "*.xlsx"))
    
    if not archivos:
        print("❌ No hay archivos .xlsx")
        return

    print(f"📂 Procesando {len(archivos)} archivos...")

    for archivo in archivos:
        nombre_archivo = os.path.basename(archivo)
        print(f"\n📄 {nombre_archivo}")
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            anio_match = re.search(r'20\d{2}', nombre_archivo)
            anio = anio_match.group(0) if anio_match else "2024"

            for sheet_name in wb.sheetnames:
                if any(x in sheet_name.lower() for x in HOJAS_IGNORAR): continue
                ws = wb[sheet_name]
                
                # 1. Encontrar columna RUT (El ancla)
                idx_rut = buscar_columna_rut(ws)
                if idx_rut == -1: continue # Hoja sin RUTs, saltar

                # 2. Encontrar fila de inicio
                fila_data = encontrar_inicio_datos(ws, idx_rut)
                
                print(f"   ✅ Hoja '{sheet_name}': RUT en Col {idx_rut}")

                # 3. Procesar filas
                count = 0
                for row in ws.iter_rows(min_row=fila_data):
                    if idx_rut >= len(row): continue
                    
                    val_rut = str(row[idx_rut].value).strip() if row[idx_rut].value else ""
                    if not es_rut_chileno(val_rut): continue

                    # --- ESTRATEGIA HÍBRIDA PARA EL NOMBRE ---
                    val_nombre = ""
                    idx_nombre_usado = -1 # Para saber qué columna saltar al leer obs
                    
                    # Intentamos mirar SIEMPRE a la derecha del RUT (Columna B usualmente)
                    posible_idx_nombre = idx_rut + 1
                    
                    if posible_idx_nombre < len(row):
                        candidato = row[posible_idx_nombre].value
                        if candidato and es_nombre_valido(candidato):
                            val_nombre = str(candidato).strip()
                            idx_nombre_usado = posible_idx_nombre
                        else:
                            # Si falla (ej: tiene numeros), intentamos RUT+2 (a veces hay columna vacia entre medio)
                            if posible_idx_nombre + 1 < len(row):
                                candidato_2 = row[posible_idx_nombre + 1].value
                                if candidato_2 and es_nombre_valido(candidato_2):
                                    val_nombre = str(candidato_2).strip()
                                    idx_nombre_usado = posible_idx_nombre + 1

                    # Si aun asi no hay nombre, lo dejamos vacio ("") pero procesamos igual
                    
                    # --- ESCANEAR ALERTAS ---
                    rojos = 0
                    amarillos = 0
                    obs_list = []

                    for i, celda in enumerate(row):
                        # Saltamos la columna del RUT y la columna donde ENCONTRAMOS el nombre
                        if i == idx_rut or i == idx_nombre_usado: continue 
                        
                        riesgo = obtener_riesgo_color(celda)
                        if riesgo == 3: rojos += 1
                        if riesgo == 2: amarillos += 1
                        
                        # Extraer texto de observaciones
                        # Solo si tiene color o es texto largo (para evitar sacar notas tipo "4.5")
                        txt = str(celda.value).strip() if celda.value else ""
                        es_texto_largo = len(txt) > 10 and not any(char.isdigit() for char in txt)
                        
                        if (riesgo > 0 or es_texto_largo) and txt:
                             if txt.lower() not in ['si', 'no', 'n/a'] and "@" not in txt:
                                obs_list.append(txt)

                    data_final.append({
                        'origen': nombre_archivo,
                        'anio': anio,
                        'tutor': sheet_name,
                        'rut': val_rut,
                        'nombre': val_nombre, 
                        'cant_rojos': rojos,
                        'cant_amarillos': amarillos,
                        'observaciones': " || ".join(set(obs_list))
                    })
                    count += 1
                
                if count > 0:
                    print(f"      -> {count} reg.")

        except Exception as e:
            print(f"   ❌ Error: {e}")

    if data_final:
        df = pd.DataFrame(data_final)
        print(f"\n✨ FIN. Total: {len(df)} registros.")
        print(df[['rut', 'nombre', 'observaciones']].head(10)) # Previsualizar nombres
        df.to_csv(ARCHIVO_SALIDA, index=False, encoding='utf-8-sig')
    else:
        print("\n⚠️ No se generaron datos.")

if __name__ == "__main__":
    procesar_excels()