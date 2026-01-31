import pandas as pd
import openpyxl
import glob
import os
import re
import warnings

warnings.filterwarnings('ignore', category=UserWarning)

# --- CONFIGURACIÓN ---
CARPETA_DATOS = "raw_data"
ARCHIVO_SALIDA = "bitacora_estructurada_final.csv"

HOJAS_IGNORAR = [
    "instrucc", "resumen", "hoja1", "base", "ejemplo", "decálogo", "decalogo", 
    "ruta", "contacto", "protocolo", "derivación", "actividades", "tabla", "validac", "listado"
]

# Mapeo de Siglas a Nombres Completos (Ajusta si me equivoqué en alguna)
MAPA_CARRERAS = {
    "ARQ": "Arquitectura",
    "BACH": "Bachillerato en Ciencias",
    "EDI": "Diseño Industrial",
    "INCO": "Ingeniería en Construcción",
    "ICO": "Ingeniería Comercial",
    "CPA": "Contador Público y Auditor",
    "DER": "Derecho", # Vi un DER 2025.xlsx
    "IES": "Ingeniería Estadística", # Asumo IES, ajusta si es otra ingeniería
    "TS": "Trabajo Social",
    "ICI": "Ingeniería Civil Industrial",
    "IECI": "Ingeniería de Ejecución en Computación e Informática",
    "ICINF": "Ingeniería Civil en Informática"
}

MAPA_COLUMNAS_BASE = {
    "Teléfono": ["telefono", "fono", "celular", "movil", "contact"],
    "Correo": ["correo", "mail", "email"],
    "Lugar Procedencia": ["procedencia", "lugar", "ciudad", "origen", "comuna"],
    "Grupo Familiar": ["familiar", "familia", "vive con"],
    "Beneficio": ["beneficio", "gratuidad", "beca", "arancel", "financiamiento"],
    "Observaciones": ["observacion", "observaciones", "comentario", "detalle", "info adicional"]
}

def limpiar_texto(val):
    if not val: return ""
    return str(val).replace('\n', ' ').strip()

def es_rut_chileno(texto):
    if not texto: return False
    limpio = str(texto).replace('.', '').strip().upper()
    return len(limpio) > 5 and any(char.isdigit() for char in limpio) and '-' in limpio

def detectar_riesgo_etiqueta(celda):
    if celda.fill and celda.fill.start_color:
        try:
            if not celda.fill.start_color.index: return ""
            c = str(celda.fill.start_color.index).upper()
            if len(c) < 3: return "" 
            if c in ['FFFF0000', '00FF0000', 'RED', 'FFFFC7CE', '3']: return "[ROJO] "
            if c in ['FFFFFF00', 'YELLOW', 'FFFFEB9C', '2']: return "[AMARILLO] "
        except:
            pass
    return ""

def escanear_ubicacion_datos(ws):
    max_filas = 50
    scores_rut = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_filas, values_only=True)):
        for col_idx, valor in enumerate(row):
            if es_rut_chileno(str(valor)):
                scores_rut[col_idx] = scores_rut.get(col_idx, 0) + 1
    
    if not scores_rut: return -1, -1
    idx_rut = max(scores_rut, key=scores_rut.get)
    fila_inicio = -1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_filas, values_only=True)):
        if es_rut_chileno(str(row[idx_rut])):
            fila_inicio = i + 1
            break
    return idx_rut, fila_inicio

def generar_mapa_columnas(ws, fila_datos, idx_rut):
    mapa = {}
    fila_header = fila_datos - 1
    if fila_header < 1: return { 'rut': idx_rut }

    row_cells = list(ws.iter_rows(min_row=fila_header, max_row=fila_header, max_col=40))[0]
    for i, cell in enumerate(row_cells):
        val = str(cell.value).lower().strip() if cell.value else ""
        if i == idx_rut:
            mapa['rut'] = i
            continue
        
        encontrado = False
        for nombre_meta, sinonimos in MAPA_COLUMNAS_BASE.items():
            if any(s in val for s in sinonimos) and "alerta" not in val:
                mapa[nombre_meta] = i
                encontrado = True
                break
        if encontrado: continue

        match_num = re.search(r'(\d+)[°º]?\s*alerta|alerta\s*(\d+)', val)
        if match_num:
            num = match_num.group(1) if match_num.group(1) else match_num.group(2)
            mapa[f"Alerta {num}"] = i
            continue

        if "espontanea" in val or "espontánea" in val:
            if "motivo" in val: mapa["Alerta Espontánea Motivo"] = i
            elif "detalle" in val or "caso" in val: mapa["Alerta Espontánea Detalle"] = i
            else: 
                if "Alerta Espontánea Motivo" not in mapa: mapa["Alerta Espontánea Motivo"] = i

    if 'nombre' not in mapa and 'rut' in mapa:
         idx_nombre_cand = mapa['rut'] + 1
         if idx_nombre_cand not in mapa.values(): mapa['nombre'] = idx_nombre_cand

    return mapa

def procesar_excels():
    data_global = []
    archivos = glob.glob(os.path.join(CARPETA_DATOS, "*.xlsx"))
    
    print(f"📂 Procesando {len(archivos)} archivos...")

    for archivo in archivos:
        nombre_archivo = os.path.basename(archivo)
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            
            # --- DETECCIÓN DE CARRERA Y AÑO ---
            nombre_upper = nombre_archivo.upper()
            
            # 1. Año
            anio_match = re.search(r'20\d{2}', nombre_archivo)
            anio = anio_match.group(0) if anio_match else "2024"
            
            # 2. Carrera
            carrera = "Otra"
            for sigla, nombre_completo in MAPA_CARRERAS.items():
                # Buscamos la sigla (ej: ARQ) rodeada de caracteres no alfabéticos o inicio/fin
                # O simplemente si está en el string (más flexible)
                if sigla in nombre_upper:
                    carrera = nombre_completo
                    break
            
            for sheet_name in wb.sheetnames:
                if any(ign in sheet_name.lower() for ign in HOJAS_IGNORAR): continue
                ws = wb[sheet_name]
                
                idx_rut, fila_datos = escanear_ubicacion_datos(ws)
                if idx_rut == -1: continue

                mapa = generar_mapa_columnas(ws, fila_datos, idx_rut)
                cols_ocupadas = {v: k for k, v in mapa.items()}
                
                for row in ws.iter_rows(min_row=fila_datos):
                    if idx_rut >= len(row): continue
                    celda_rut = row[idx_rut]
                    val_rut = limpiar_texto(celda_rut.value)
                    if not es_rut_chileno(val_rut): continue

                    # ITEM BASE CON CARRERA PRIMERO
                    item = {
                        "Origen": nombre_archivo, 
                        "Carrera": carrera,  # <--- NUEVO CAMPO
                        "Anio": anio, 
                        "Tutor": sheet_name,
                        "rut": val_rut, "nombre": ""
                    }
                    
                    if 'nombre' in mapa and mapa['nombre'] < len(row):
                        val_nom = limpiar_texto(row[mapa['nombre']].value)
                        if not any(c.isdigit() for c in val_nom): item['nombre'] = val_nom

                    for campo, idx in mapa.items():
                        if campo in ['rut', 'nombre']: continue
                        if idx < len(row):
                            celda = row[idx]
                            txt = limpiar_texto(celda.value)
                            tag = detectar_riesgo_etiqueta(celda)
                            if txt or tag: item[campo] = f"{tag}{txt}"

                    obs_extra = []
                    for i, celda in enumerate(row):
                        if i in cols_ocupadas: continue
                        tag = detectar_riesgo_etiqueta(celda)
                        txt = limpiar_texto(celda.value)
                        es_texto_largo = len(txt) > 15 and "si" not in txt.lower() and "no" not in txt.lower()
                        
                        if tag or (es_texto_largo and i > idx_rut):
                            contenido = f"{tag}{txt}"
                            if contenido.strip(): obs_extra.append(contenido)
                    
                    if obs_extra:
                        existente = item.get("Observaciones", "")
                        nuevo = " || ".join(obs_extra)
                        item["Observaciones"] = f"{existente} || {nuevo}" if existente else nuevo

                    data_global.append(item)

        except Exception as e:
            print(f"❌ Error en {nombre_archivo}: {e}")

    if data_global:
        df = pd.DataFrame(data_global)
        
        # --- ORDENAMIENTO FINAL DE COLUMNAS ---
        cols_fijas = ["Origen", "Carrera", "Anio", "Tutor", "rut", "nombre", "Teléfono", "Correo", 
                      "Lugar Procedencia", "Grupo Familiar", "Beneficio", "Observaciones"]
        
        cols_dinamicas = [c for c in df.columns if c not in cols_fijas]
        
        def sort_alertas(c):
            nums = re.findall(r'\d+', c)
            return int(nums[0]) if nums else 999
            
        cols_alertas = sorted([c for c in cols_dinamicas if "Alerta" in c and "Espontánea" not in c], key=sort_alertas)
        cols_resto = [c for c in cols_dinamicas if c not in cols_alertas]
        
        df = df[cols_fijas + cols_alertas + cols_resto]
        df.fillna("", inplace=True)
        
        print(f"\n✨ PROCESO TERMINADO. Total: {len(df)} registros.")
        df.to_csv(ARCHIVO_SALIDA, index=False, encoding='utf-8-sig')
    else:
        print("⚠️ No se encontraron datos.")

if __name__ == "__main__":
    procesar_excels()